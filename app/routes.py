from fastapi import APIRouter, HTTPException, Response
import sqlite3
import random
import io
import os
import zipfile
from urllib.parse import quote
from pydantic import BaseModel, Field
from app.database import get_connection

import openpyxl
from openpyxl.styles import Font as XFont, Alignment as XAlign, PatternFill as XFill, Border as XBorder, Side as XSide

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def get_pdf_font_name():
    fonts = [
        ("Arial", "C:\\Windows\\Fonts\\arial.ttf", "C:\\Windows\\Fonts\\arialbd.ttf"),
        ("Calibri", "C:\\Windows\\Fonts\\calibri.ttf", "C:\\Windows\\Fonts\\calibrib.ttf"),
        ("SegoeUI", "C:\\Windows\\Fonts\\segoeui.ttf", "C:\\Windows\\Fonts\\segoeuib.ttf"),
    ]
    for font_name, reg_path, bold_path in fonts:
        if os.path.exists(reg_path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, reg_path))
                if os.path.exists(bold_path):
                    pdfmetrics.registerFont(TTFont(f"{font_name}-Bold", bold_path))
                return font_name, f"{font_name}-Bold" if os.path.exists(bold_path) else font_name
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"

PDF_FONT, PDF_FONT_BOLD = get_pdf_font_name()


router = APIRouter(prefix="/api")

DAYS = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
HOURS = range(1, 9)
DOUBLE_BLOCK_STARTS = [1, 3, 5, 7]
MAX_GENERATION_ATTEMPTS = 30


class TeacherCreate(BaseModel):
    name: str = Field(min_length=2)
    branch: str = Field(min_length=2)
    max_weekly_hours: int = Field(default=24, ge=1, le=40)


class LessonCreate(BaseModel):
    name: str = Field(min_length=2)
    weekly_hours: int = Field(ge=1, le=10)
    level_group: str = Field(description="İlkokul veya Ortaokul-Lise")


class ClassCreate(BaseModel):
    grade: int = Field(ge=1, le=12)
    section: str = Field(min_length=1, max_length=1)


class RoomCreate(BaseModel):
    name: str = Field(min_length=2)
    room_type: str = Field(min_length=2)


class TeacherUnavailableCreate(BaseModel):
    teacher_id: int = Field(ge=1)
    day: str = Field(min_length=2)
    hour: int = Field(ge=1, le=8)


class UserCreate(BaseModel):
    name: str = Field(min_length=2)
    role: str = Field(min_length=2)


class ScheduleMove(BaseModel):
    class_id: int = Field(ge=1)
    from_day: str = Field(min_length=2)
    from_hour: int = Field(ge=1, le=8)
    to_day: str = Field(min_length=2)
    to_hour: int = Field(ge=1, le=8)


@router.get("/health")
def health_check():
    return {"status": "ok"}


@router.post("/teachers")
def create_teacher(payload: TeacherCreate):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO teachers (name, branch, max_weekly_hours) VALUES (?, ?, ?)",
        (payload.name, payload.branch, payload.max_weekly_hours),
    )
    conn.commit()
    teacher_id = cursor.lastrowid
    conn.close()
    return {"id": teacher_id, **payload.model_dump()}


@router.get("/teachers")
def list_teachers():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM teachers ORDER BY id").fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.delete("/teachers/{teacher_id}")
def delete_teacher(teacher_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    in_schedule = cursor.execute(
        "SELECT 1 FROM schedules WHERE teacher_id = ? LIMIT 1",
        (teacher_id,),
    ).fetchone()

    if in_schedule:
        conn.close()
        raise HTTPException(status_code=400, detail="Programda kullanılan öğretmen silinemez")

    cursor.execute("DELETE FROM teachers WHERE id = ?", (teacher_id,))
    conn.commit()
    deleted = cursor.rowcount
    conn.close()

    if deleted == 0:
        raise HTTPException(status_code=404, detail="Öğretmen bulunamadı")

    return {"message": "Öğretmen silindi"}


@router.post("/lessons")
def create_lesson(payload: LessonCreate):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO lessons (name, weekly_hours, level_group) VALUES (?, ?, ?)",
            (payload.name, payload.weekly_hours, payload.level_group),
        )
        conn.commit()
        lesson_id = cursor.lastrowid
    except sqlite3.IntegrityError as exc:
        conn.close()
        raise HTTPException(status_code=400, detail="Bu ders zaten kayıtlı") from exc
    conn.close()
    return {"id": lesson_id, **payload.model_dump()}


@router.get("/lessons")
def list_lessons():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM lessons ORDER BY id").fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.delete("/lessons/{lesson_id}")
def delete_lesson(lesson_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    in_schedule = cursor.execute(
        "SELECT 1 FROM schedules WHERE lesson_id = ? LIMIT 1",
        (lesson_id,),
    ).fetchone()

    if in_schedule:
        conn.close()
        raise HTTPException(status_code=400, detail="Programda kullanılan ders silinemez")

    cursor.execute("DELETE FROM lessons WHERE id = ?", (lesson_id,))
    conn.commit()
    deleted = cursor.rowcount
    conn.close()

    if deleted == 0:
        raise HTTPException(status_code=404, detail="Ders bulunamadı")

    return {"message": "Ders silindi"}


@router.post("/classes")
def create_class(payload: ClassCreate):
    class_name = f"{payload.grade}-{payload.section.upper()}"

    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            "INSERT INTO classes (grade, section, name) VALUES (?, ?, ?)",
            (payload.grade, payload.section.upper(), class_name),
        )
        conn.commit()
    except Exception as exc:
        conn.close()
        raise HTTPException(status_code=400, detail="Bu sınıf zaten kayıtlı olabilir") from exc

    class_id = cursor.lastrowid
    conn.close()
    return {"id": class_id, "grade": payload.grade, "section": payload.section.upper(), "name": class_name}


@router.get("/classes")
def list_classes():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM classes ORDER BY grade, section").fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.post("/rooms")
def create_room(payload: RoomCreate):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO rooms (name, room_type) VALUES (?, ?)",
            (payload.name, payload.room_type),
        )
        conn.commit()
        room_id = cursor.lastrowid
    except sqlite3.IntegrityError as exc:
        conn.close()
        raise HTTPException(status_code=400, detail="Bu derslik zaten kayıtlı") from exc
    conn.close()
    return {"id": room_id, **payload.model_dump()}


@router.get("/rooms")
def list_rooms():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM rooms ORDER BY room_type, name").fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.delete("/rooms/{room_id}")
def delete_room(room_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    in_schedule = cursor.execute(
        "SELECT 1 FROM schedules WHERE room_id = ? LIMIT 1",
        (room_id,),
    ).fetchone()

    if in_schedule:
        conn.close()
        raise HTTPException(status_code=400, detail="Programda kullanılan derslik silinemez")

    cursor.execute("DELETE FROM rooms WHERE id = ?", (room_id,))
    conn.commit()
    deleted = cursor.rowcount
    conn.close()

    if deleted == 0:
        raise HTTPException(status_code=404, detail="Derslik bulunamadı")

    return {"message": "Derslik silindi"}


@router.post("/teacher-unavailability")
def create_teacher_unavailability(payload: TeacherUnavailableCreate):
    conn = get_connection()
    cursor = conn.cursor()
    teacher = cursor.execute("SELECT id FROM teachers WHERE id = ?", (payload.teacher_id,)).fetchone()
    if teacher is None:
        conn.close()
        raise HTTPException(status_code=404, detail="Öğretmen bulunamadı")

    try:
        cursor.execute(
            "INSERT INTO teacher_unavailability (teacher_id, day, hour) VALUES (?, ?, ?)",
            (payload.teacher_id, payload.day, payload.hour),
        )
        conn.commit()
        item_id = cursor.lastrowid
    except sqlite3.IntegrityError as exc:
        conn.close()
        raise HTTPException(status_code=400, detail="Bu uygun olmayan saat zaten kayıtlı") from exc
    conn.close()
    return {"id": item_id, **payload.model_dump()}


@router.get("/teacher-unavailability")
def list_teacher_unavailability():
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            teacher_unavailability.id,
            teacher_unavailability.teacher_id,
            teachers.name AS teacher_name,
            teachers.branch,
            teacher_unavailability.day,
            teacher_unavailability.hour
        FROM teacher_unavailability
        JOIN teachers ON teachers.id = teacher_unavailability.teacher_id
        ORDER BY teachers.name, teacher_unavailability.day, teacher_unavailability.hour
        """
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.delete("/teacher-unavailability/{item_id}")
def delete_teacher_unavailability(item_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM teacher_unavailability WHERE id = ?", (item_id,))
    conn.commit()
    deleted = cursor.rowcount
    conn.close()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    return {"message": "Uygun olmayan saat silindi"}


@router.post("/users")
def create_user(payload: UserCreate):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO users (name, role) VALUES (?, ?)",
        (payload.name, payload.role),
    )
    conn.commit()
    user_id = cursor.lastrowid
    conn.close()
    return {"id": user_id, **payload.model_dump()}


@router.get("/users")
def list_users():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM users ORDER BY role, name").fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.delete("/users/{user_id}")
def delete_user(user_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    deleted = cursor.rowcount
    conn.close()
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return {"message": "Kullanıcı silindi"}


@router.get("/teacher-loads")
def list_teacher_loads():
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            teachers.id,
            teachers.name,
            teachers.branch,
            teachers.max_weekly_hours,
            COUNT(schedules.id) AS assigned_hours
        FROM teachers
        LEFT JOIN schedules ON schedules.teacher_id = teachers.id
        GROUP BY teachers.id, teachers.name, teachers.branch, teachers.max_weekly_hours
        ORDER BY teachers.branch, teachers.name
        """
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.get("/class-status")
def list_class_status():
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            classes.id,
            classes.grade,
            classes.section,
            classes.name,
            COUNT(schedules.id) AS assigned_hours
        FROM classes
        LEFT JOIN schedules ON schedules.class_id = classes.id
        GROUP BY classes.id, classes.grade, classes.section, classes.name
        ORDER BY classes.grade, classes.section
        """
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.get("/lesson-summary")
def list_lesson_summary():
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            lessons.id,
            lessons.name,
            lessons.weekly_hours,
            lessons.level_group,
            COUNT(schedules.id) AS assigned_hours
        FROM lessons
        LEFT JOIN schedules ON schedules.lesson_id = lessons.id
        GROUP BY lessons.id, lessons.name, lessons.weekly_hours, lessons.level_group
        ORDER BY lessons.level_group, lessons.name
        """
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.get("/teacher-schedule/{teacher_id}")
def get_teacher_schedule(teacher_id: int):
    conn = get_connection()
    teacher = conn.execute(
        "SELECT * FROM teachers WHERE id = ?",
        (teacher_id,),
    ).fetchone()

    if teacher is None:
        conn.close()
        raise HTTPException(status_code=404, detail="Öğretmen bulunamadı")

    rows = conn.execute(
        """
        SELECT
            schedules.day,
            schedules.hour,
            classes.name AS class_name,
            lessons.name AS lesson_name,
            teachers.name AS teacher_name,
            rooms.name AS room_name
        FROM schedules
        JOIN classes ON classes.id = schedules.class_id
        JOIN lessons ON lessons.id = schedules.lesson_id
        JOIN teachers ON teachers.id = schedules.teacher_id
        LEFT JOIN rooms ON rooms.id = schedules.room_id
        WHERE schedules.teacher_id = ?
        ORDER BY
            CASE schedules.day
                WHEN 'Pazartesi' THEN 1
                WHEN 'Salı' THEN 2
                WHEN 'Çarşamba' THEN 3
                WHEN 'Perşembe' THEN 4
                WHEN 'Cuma' THEN 5
            END,
            schedules.hour
        """,
        (teacher_id,),
    ).fetchall()
    conn.close()
    return {"teacher": dict(teacher), "schedule": [dict(row) for row in rows]}


@router.get("/validation-report")
def get_validation_report():
    conn = get_connection()
    class_status = [dict(row) for row in conn.execute(
        """
        SELECT classes.name, COUNT(schedules.id) AS assigned_hours
        FROM classes
        LEFT JOIN schedules ON schedules.class_id = classes.id
        GROUP BY classes.id, classes.name
        HAVING assigned_hours != 40
        ORDER BY classes.name
        """
    )]
    class_conflicts = [dict(row) for row in conn.execute(
        """
        SELECT classes.name AS class_name, schedules.day, schedules.hour, COUNT(*) AS total
        FROM schedules
        JOIN classes ON classes.id = schedules.class_id
        GROUP BY schedules.class_id, schedules.day, schedules.hour
        HAVING total > 1
        """
    )]
    teacher_conflicts = [dict(row) for row in conn.execute(
        """
        SELECT teachers.name AS teacher_name, schedules.day, schedules.hour, COUNT(*) AS total
        FROM schedules
        JOIN teachers ON teachers.id = schedules.teacher_id
        GROUP BY schedules.teacher_id, schedules.day, schedules.hour
        HAVING total > 1
        """
    )]
    teacher_overload = [dict(row) for row in conn.execute(
        """
        SELECT teachers.name, teachers.branch, COUNT(schedules.id) AS assigned_hours, teachers.max_weekly_hours
        FROM teachers
        JOIN schedules ON schedules.teacher_id = teachers.id
        GROUP BY teachers.id, teachers.name, teachers.branch, teachers.max_weekly_hours
        HAVING assigned_hours > teachers.max_weekly_hours
        ORDER BY teachers.branch, teachers.name
        """
    )]
    room_conflicts = [dict(row) for row in conn.execute(
        """
        SELECT rooms.name AS room_name, schedules.day, schedules.hour, COUNT(*) AS total
        FROM schedules
        JOIN rooms ON rooms.id = schedules.room_id
        GROUP BY schedules.room_id, schedules.day, schedules.hour
        HAVING total > 1
        """
    )]
    teacher_unavailable_conflicts = [dict(row) for row in conn.execute(
        """
        SELECT teachers.name AS teacher_name, schedules.day, schedules.hour
        FROM schedules
        JOIN teachers ON teachers.id = schedules.teacher_id
        JOIN teacher_unavailability
            ON teacher_unavailability.teacher_id = schedules.teacher_id
            AND teacher_unavailability.day = schedules.day
            AND teacher_unavailability.hour = schedules.hour
        ORDER BY teachers.name, schedules.day, schedules.hour
        """
    )]
    total_hours = conn.execute("SELECT COUNT(*) AS total FROM schedules").fetchone()["total"]
    conn.close()
    return {
        "total_hours": total_hours,
        "class_status": class_status,
        "class_conflicts": class_conflicts,
        "teacher_conflicts": teacher_conflicts,
        "teacher_overload": teacher_overload,
        "room_conflicts": room_conflicts,
        "teacher_unavailable_conflicts": teacher_unavailable_conflicts,
    }


@router.post("/generate-schedule/{class_id}")
def generate_schedule(class_id: int):
    conn = get_connection()
    cursor = conn.cursor()

    try:
        placed_count = try_generate_schedule_for_class(conn, cursor, class_id)
        conn.commit()
    except ValueError as exc:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    conn.close()

    message = "Program oluşturuldu"
    if placed_count < 40:
        message = f"Program olusturuldu ancak haftalik ders saati 40 degil: {placed_count}"

    return {
        "message": message,
        "class_id": class_id,
        "placed_hours": placed_count,
    }


@router.post("/generate-schedules")
def generate_all_schedules():
    conn = get_connection()
    cursor = conn.cursor()
    classes = cursor.execute("SELECT * FROM classes ORDER BY grade, section").fetchall()

    if not classes:
        conn.close()
        raise HTTPException(status_code=400, detail="Önce sınıf eklemelisin")

    try:
        last_error = None
        results = []

        for attempt in range(MAX_GENERATION_ATTEMPTS):
            rng = random.Random()
            cursor.execute("DELETE FROM schedules")
            results = []
            class_order = list(classes)
            if attempt > 0:
                rng.shuffle(class_order)

            try:
                for class_row in class_order:
                    placed_count = generate_schedule_for_class(
                        conn,
                        cursor,
                        class_row["id"],
                        clear_existing=False,
                        rng=rng,
                    )
                    results.append({
                        "class_id": class_row["id"],
                        "class_name": class_row["name"],
                        "placed_hours": placed_count,
                    })
                break
            except ValueError as exc:
                last_error = exc
        else:
            raise ValueError(f"Kurallara uygun program bulunamadi: {last_error}")

        conn.commit()
    except ValueError as exc:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    total_hours = sum(item["placed_hours"] for item in results)
    incomplete_classes = [item for item in results if item["placed_hours"] < 40]
    conn.close()

    message = "Tüm sınıflar için program oluşturuldu"
    if incomplete_classes:
        message = f"Program olusturuldu ancak {len(incomplete_classes)} sınıf 40 saatin altında"

    return {
        "message": message,
        "class_count": len(results),
        "total_hours": total_hours,
        "incomplete_classes": incomplete_classes,
    }


def try_generate_schedule_for_class(conn, cursor, class_id: int) -> int:
    last_error = None

    for _attempt in range(MAX_GENERATION_ATTEMPTS):
        rng = random.Random()
        cursor.execute("DELETE FROM schedules WHERE class_id = ?", (class_id,))

        try:
            return generate_schedule_for_class(
                conn,
                cursor,
                class_id,
                clear_existing=False,
                rng=rng,
            )
        except ValueError as exc:
            last_error = exc
            cursor.execute("DELETE FROM schedules WHERE class_id = ?", (class_id,))

    raise ValueError(f"Kurallara uygun program bulunamadi: {last_error}")


def generate_schedule_for_class(conn, cursor, class_id: int, clear_existing: bool, rng) -> int:
    class_row = cursor.execute("SELECT * FROM classes WHERE id = ?", (class_id,)).fetchone()
    if class_row is None:
        raise ValueError("Sınıf bulunamadı")

    level_group = get_level_group_for_grade(class_row["grade"])
    lessons = get_lessons_for_level(conn, level_group)
    if not lessons:
        raise ValueError(f"{level_group} icin önce ders eklemelisin")

    total_lesson_hours = sum(lesson["weekly_hours"] for lesson in lessons)
    if total_lesson_hours > 40:
        raise ValueError(f"Ders toplam saati 40'ı geçemez. Şu an: {total_lesson_hours}")

    if clear_existing:
        cursor.execute("DELETE FROM schedules WHERE class_id = ?", (class_id,))

    placed_count = 0
    for lesson in lessons:
        remaining_hours = lesson["weekly_hours"]

        while remaining_hours > 0:
            block_size = 2 if remaining_hours >= 2 else 1
            placement = find_placement_for_lesson(conn, class_id, lesson, block_size, rng)

            if placement is None:
                raise ValueError(
                    f"{class_row['name']} sınıfı için {lesson['name']} dersine uygun öğretmen veya saat bulunamadı"
                )

            teacher, day, start_hour, room = placement

            for offset in range(block_size):
                cursor.execute(
                    """
                    INSERT INTO schedules (class_id, lesson_id, teacher_id, room_id, day, hour)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (class_id, lesson["id"], teacher["id"], room["id"] if room else None, day, start_hour + offset),
                )
                placed_count += 1

            remaining_hours -= block_size

    return placed_count


def get_level_group_for_grade(grade: int) -> str:
    if 1 <= grade <= 4:
        return "İlkokul"
    return "Ortaokul-Lise"


def get_lessons_for_level(conn, level_group: str):
    return conn.execute(
        """
        SELECT * FROM lessons
        WHERE level_group IN (?, 'Tüm')
        ORDER BY weekly_hours DESC, name
        """,
        (level_group,),
    ).fetchall()


@router.get("/schedule/{class_id}")
def get_schedule(class_id: int):
    conn = get_connection()
    rows = conn.execute(
        """
        SELECT
            schedules.day,
            schedules.hour,
            classes.name AS class_name,
            lessons.name AS lesson_name,
            teachers.name AS teacher_name,
            rooms.name AS room_name
        FROM schedules
        JOIN classes ON classes.id = schedules.class_id
        JOIN lessons ON lessons.id = schedules.lesson_id
        JOIN teachers ON teachers.id = schedules.teacher_id
        LEFT JOIN rooms ON rooms.id = schedules.room_id
        WHERE schedules.class_id = ?
        ORDER BY
            CASE schedules.day
                WHEN 'Pazartesi' THEN 1
                WHEN 'Salı' THEN 2
                WHEN 'Çarşamba' THEN 3
                WHEN 'Perşembe' THEN 4
                WHEN 'Cuma' THEN 5
            END,
            schedules.hour
        """,
        (class_id,),
    ).fetchall()
    conn.close()
    return [dict(row) for row in rows]


@router.post("/schedule/move")
def move_schedule_item(payload: ScheduleMove):
    conn = get_connection()
    cursor = conn.cursor()
    source = cursor.execute(
        """
        SELECT schedules.*, lessons.name AS lesson_name
        FROM schedules
        JOIN lessons ON lessons.id = schedules.lesson_id
        WHERE schedules.class_id = ? AND schedules.day = ? AND schedules.hour = ?
        """,
        (payload.class_id, payload.from_day, payload.from_hour),
    ).fetchone()

    if source is None:
        conn.close()
        raise HTTPException(status_code=404, detail="Taşınacak ders bulunamadı")

    target = cursor.execute(
        """
        SELECT schedules.*, lessons.name AS lesson_name
        FROM schedules
        JOIN lessons ON lessons.id = schedules.lesson_id
        WHERE schedules.class_id = ? AND schedules.day = ? AND schedules.hour = ?
        """,
        (payload.class_id, payload.to_day, payload.to_hour),
    ).fetchone()

    if target and target["id"] == source["id"]:
        conn.close()
        return {"message": "Ders zaten seçilen saatte"}

    try:
        moved_items = [source]
        if target:
            moved_items.append(target)

        for item in moved_items:
            cursor.execute("DELETE FROM schedules WHERE id = ?", (item["id"],))

        source_room = choose_room_for_manual_move(conn, source, payload.to_day, payload.to_hour)
        if not can_place_block(conn, payload.class_id, source["lesson_id"], source["teacher_id"], payload.to_day, payload.to_hour, 1, source_room):
            raise ValueError("Hedef saat kurallara uygun değil")

        cursor.execute(
            """
            INSERT INTO schedules (class_id, lesson_id, teacher_id, room_id, day, hour)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (payload.class_id, source["lesson_id"], source["teacher_id"], source_room["id"] if source_room else None, payload.to_day, payload.to_hour),
        )

        if target:
            target_room = choose_room_for_manual_move(conn, target, payload.from_day, payload.from_hour)
            if not can_place_block(conn, payload.class_id, target["lesson_id"], target["teacher_id"], payload.from_day, payload.from_hour, 1, target_room):
                raise ValueError("Yer değiştirilecek ders kaynak saate uygun değil")

            cursor.execute(
                """
                INSERT INTO schedules (class_id, lesson_id, teacher_id, room_id, day, hour)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (payload.class_id, target["lesson_id"], target["teacher_id"], target_room["id"] if target_room else None, payload.from_day, payload.from_hour),
            )

        conn.commit()
    except ValueError as exc:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    conn.close()
    return {"message": "Ders taşındı" if not target else "Derslerin yeri değiştirildi"}


@router.get("/quality-report")
def get_quality_report():
    conn = get_connection()
    teacher_gaps = 0
    teacher_heavy_days = 0
    lesson_balance_warnings = 0

    teachers = conn.execute("SELECT id FROM teachers").fetchall()
    for teacher in teachers:
        for day in DAYS:
            hours = [
                row["hour"]
                for row in conn.execute(
                    "SELECT hour FROM schedules WHERE teacher_id = ? AND day = ? ORDER BY hour",
                    (teacher["id"], day),
                ).fetchall()
            ]
            if len(hours) >= 2:
                teacher_gaps += max(0, max(hours) - min(hours) + 1 - len(hours))
            if len(hours) > 6:
                teacher_heavy_days += 1

    class_lessons = conn.execute(
        """
        SELECT class_id, lesson_id, day, COUNT(*) AS total
        FROM schedules
        GROUP BY class_id, lesson_id, day
        HAVING total > 2
        """
    ).fetchall()
    lesson_balance_warnings = len(class_lessons)

    penalty = min(60, teacher_gaps * 2 + teacher_heavy_days * 3 + lesson_balance_warnings * 5)
    score = max(0, 100 - penalty)
    conn.close()

    return {
        "score": score,
        "teacher_gaps": teacher_gaps,
        "teacher_heavy_days": teacher_heavy_days,
        "lesson_balance_warnings": lesson_balance_warnings,
    }


def build_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f'attachment; filename="{encoded}"; filename*=UTF-8\'\'{encoded}'


@router.get("/export/class/{class_id}.xlsx")
def export_class_xlsx(class_id: int):
    schedule = get_schedule(class_id)
    conn = get_connection()
    class_row = conn.execute("SELECT * FROM classes WHERE id = ?", (class_id,)).fetchone()
    conn.close()
    if class_row is None:
        raise HTTPException(status_code=404, detail="Sınıf bulunamadı")
    title = f"{class_row['name']} SINIFI HAFTALIK DERS PROGRAMI"
    content = build_formatted_xlsx(title, schedule, primary_label="lesson_name")
    filename = f"{class_row['name']}_program.xlsx"
    return Response(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": build_content_disposition(filename)},
    )


@router.get("/export/teacher/{teacher_id}.xlsx")
def export_teacher_xlsx(teacher_id: int):
    result = get_teacher_schedule(teacher_id)
    title = f"{result['teacher']['name']} ÖĞRETMEN HAFTALIK DERS PROGRAMI"
    content = build_formatted_xlsx(title, result["schedule"], primary_label="class_name")
    filename = f"{result['teacher']['name'].replace(' ', '_')}_program.xlsx"
    return Response(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": build_content_disposition(filename)},
    )


@router.get("/export/class/{class_id}.pdf")
def export_class_pdf(class_id: int):
    schedule = get_schedule(class_id)
    conn = get_connection()
    class_row = conn.execute("SELECT * FROM classes WHERE id = ?", (class_id,)).fetchone()
    conn.close()
    if class_row is None:
        raise HTTPException(status_code=404, detail="Sınıf bulunamadı")
    title = f"{class_row['name']} SINIFI HAFTALIK DERS PROGRAMI"
    content = build_formatted_pdf(title, schedule, primary_label="lesson_name")
    filename = f"{class_row['name']}_program.pdf"
    return Response(
        content,
        media_type="application/pdf",
        headers={"Content-Disposition": build_content_disposition(filename)},
    )


@router.get("/export/teacher/{teacher_id}.pdf")
def export_teacher_pdf(teacher_id: int):
    result = get_teacher_schedule(teacher_id)
    title = f"{result['teacher']['name']} ÖĞRETMEN HAFTALIK DERS PROGRAMI"
    content = build_formatted_pdf(title, result["schedule"], primary_label="class_name")
    filename = f"{result['teacher']['name'].replace(' ', '_')}_program.pdf"
    return Response(
        content,
        media_type="application/pdf",
        headers={"Content-Disposition": build_content_disposition(filename)},
    )


@router.delete("/schedule/{class_id}")
def clear_schedule(class_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM schedules WHERE class_id = ?", (class_id,))
    conn.commit()
    deleted = cursor.rowcount
    conn.close()
    return {"message": "Program temizlendi", "deleted": deleted}


@router.delete("/schedules")
def clear_all_schedules():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM schedules")
    conn.commit()
    deleted = cursor.rowcount
    conn.close()
    return {"message": "Tüm programlar temizlendi", "deleted": deleted}


def get_teacher_weekly_load(conn, teacher_id: int, ignore_schedule_id=None) -> int:
    if ignore_schedule_id is not None:
        query = "SELECT COUNT(*) AS total FROM schedules WHERE teacher_id = ? AND id != ?"
        params = (teacher_id, ignore_schedule_id)
    else:
        query = "SELECT COUNT(*) AS total FROM schedules WHERE teacher_id = ?"
        params = (teacher_id,)
    row = conn.execute(query, params).fetchone()
    return row["total"] if row else 0


def lesson_requires_room(lesson_name: str) -> bool:
    return lesson_name in ["Beden Eğitimi", "Resim", "Görsel Sanatlar", "Müzik", "Satranç", "Fen Bilgisi"]


def choose_available_room(conn, lesson_name: str, day: str, hour: int, block_size: int, rng):
    room_type_map = {
        "Beden Eğitimi": "Beden Eğitimi",
        "Resim": "Resim",
        "Görsel Sanatlar": "Resim",
        "Müzik": "Müzik",
        "Satranç": "Satranç",
        "Fen Bilgisi": "Fen Bilgisi",
    }
    room_type = room_type_map.get(lesson_name)
    if not room_type:
        return None

    rooms = conn.execute(
        "SELECT * FROM rooms WHERE room_type = ? ORDER BY id",
        (room_type,),
    ).fetchall()

    candidate_rooms = []
    for room in rooms:
        conflict = conn.execute(
            """
            SELECT 1 FROM schedules
            WHERE room_id = ? AND day = ? AND hour BETWEEN ? AND ?
            LIMIT 1
            """,
            (room["id"], day, hour, hour + block_size - 1),
        ).fetchone()
        if not conflict:
            candidate_rooms.append(room)

    if not candidate_rooms:
        return None

    rng.shuffle(candidate_rooms)
    return candidate_rooms[0]


def can_place_block(
    conn,
    class_id: int,
    lesson_id: int,
    teacher_id: int,
    day: str,
    start_hour: int,
    block_size: int,
    room=None,
    ignore_schedule_id=None,
) -> bool:
    teacher = conn.execute(
        "SELECT * FROM teachers WHERE id = ?",
        (teacher_id,),
    ).fetchone()

    if teacher is None:
        return False

    if get_teacher_weekly_load(conn, teacher_id, ignore_schedule_id) + block_size > teacher["max_weekly_hours"]:
        return False

    unavailable = conn.execute(
        """
        SELECT 1 FROM teacher_unavailability
        WHERE teacher_id = ? AND day = ? AND hour BETWEEN ? AND ?
        LIMIT 1
        """,
        (teacher_id, day, start_hour, start_hour + block_size - 1),
    ).fetchone()

    if unavailable:
        return False

    class_conflict = conn.execute(
        """
        SELECT 1 FROM schedules
        WHERE class_id = ? AND day = ? AND hour BETWEEN ? AND ?
        LIMIT 1
        """,
        (class_id, day, start_hour, start_hour + block_size - 1),
    ).fetchone()

    if class_conflict:
        return False

    teacher_conflict = conn.execute(
        """
        SELECT 1 FROM schedules
        WHERE teacher_id = ? AND day = ? AND hour BETWEEN ? AND ?
        LIMIT 1
        """,
        (teacher_id, day, start_hour, start_hour + block_size - 1),
    ).fetchone()

    if teacher_conflict:
        return False

    if room is not None:
        room_conflict = conn.execute(
            """
            SELECT 1 FROM schedules
            WHERE room_id = ? AND day = ? AND hour BETWEEN ? AND ?
            LIMIT 1
            """,
            (room["id"], day, start_hour, start_hour + block_size - 1),
        ).fetchone()
        if room_conflict:
            return False

    return True


def find_placement_for_lesson(conn, class_id: int, lesson, block_size: int, rng):
    teachers = conn.execute(
        "SELECT * FROM teachers WHERE branch = ? ORDER BY id",
        (lesson["name"],),
    ).fetchall()

    available_teachers = []
    for teacher in teachers:
        current_load = get_teacher_weekly_load(conn, teacher["id"])
        if current_load + block_size <= teacher["max_weekly_hours"]:
            available_teachers.append((teacher, current_load))

    rng.shuffle(available_teachers)
    available_teachers.sort(key=lambda item: item[1])

    for teacher, _current_load in available_teachers:
        slot = find_available_slot(conn, class_id, lesson["id"], teacher["id"], lesson["name"], block_size, rng)
        if slot is not None:
            day, start_hour, room = slot
            return teacher, day, start_hour, room

    return None


def find_available_slot(conn, class_id: int, lesson_id: int, teacher_id: int, lesson_name: str, block_size: int, rng):
    day_offset = (class_id - 1) % len(DAYS)
    candidate_days = rotate_list(DAYS, day_offset)
    rng.shuffle(candidate_days)

    if block_size == 2:
        hour_offset = ((class_id - 1) // len(DAYS)) % len(DOUBLE_BLOCK_STARTS)
        candidate_hours = rotate_list(DOUBLE_BLOCK_STARTS, hour_offset)
    else:
        candidate_hours = list(HOURS)
    rng.shuffle(candidate_hours)

    for day in candidate_days:
        for hour in candidate_hours:
            if hour + block_size - 1 > 8:
                continue

            room = choose_available_room(conn, lesson_name, day, hour, block_size, rng)
            if lesson_requires_room(lesson_name) and room is None:
                continue
            if can_place_block(conn, class_id, lesson_id, teacher_id, day, hour, block_size, room):
                return day, hour, room

    return None


def rotate_list(items, offset: int):
    offset = offset % len(items)
    return list(items[offset:]) + list(items[:offset])


def build_formatted_xlsx(title: str, schedule: list, primary_label: str = "lesson_name") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ders Programı"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = title.upper()
    title_cell.font = XFont(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = XFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    title_cell.alignment = XAlign(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 10

    headers = ["Saat", *DAYS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = XFont(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = XFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.alignment = XAlign(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    schedule_map = {(row["day"], row["hour"]): row for row in schedule}
    thin_border = XBorder(
        left=XSide(style="thin", color="CBD5E1"),
        right=XSide(style="thin", color="CBD5E1"),
        top=XSide(style="thin", color="CBD5E1"),
        bottom=XSide(style="thin", color="CBD5E1")
    )

    for hour in HOURS:
        row_num = hour + 3
        ws.row_dimensions[row_num].height = 42

        hour_cell = ws.cell(row=row_num, column=1, value=f"{hour}. Saat")
        hour_cell.font = XFont(name="Calibri", size=11, bold=True, color="334155")
        hour_cell.fill = XFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        hour_cell.alignment = XAlign(horizontal="center", vertical="center")
        hour_cell.border = thin_border

        for col_idx, day in enumerate(DAYS, start=2):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            item = schedule_map.get((day, hour))
            if item:
                main_text = item.get(primary_label) or item.get("lesson_name") or ""
                sub_text = item.get("lesson_name") if primary_label == "class_name" else item.get("teacher_name") or ""
                room_text = f" ({item['room_name']})" if item.get("room_name") else ""

                full_text = f"{main_text}\n{sub_text}{room_text}"
                cell.value = full_text
                cell.font = XFont(name="Calibri", size=10, bold=True, color="0F172A")
                cell.alignment = XAlign(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = XFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            else:
                cell.value = "-"
                cell.font = XFont(name="Calibri", size=10, color="94A3B8")
                cell.alignment = XAlign(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, 7):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 24

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_formatted_pdf(title: str, schedule: list, primary_label: str = "lesson_name") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=30,
        rightMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    story = []

    title_style = ParagraphStyle(
        name="TitleStyle",
        fontName=PDF_FONT_BOLD,
        fontSize=16,
        leading=20,
        alignment=1,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=15
    )

    head_style = ParagraphStyle(
        name="HeadStyle",
        fontName=PDF_FONT_BOLD,
        fontSize=10,
        leading=12,
        alignment=1,
        textColor=colors.white
    )

    cell_main_style = ParagraphStyle(
        name="CellMainStyle",
        fontName=PDF_FONT_BOLD,
        fontSize=9,
        leading=11,
        alignment=1,
        textColor=colors.HexColor("#0F172A")
    )

    cell_empty_style = ParagraphStyle(
        name="CellEmptyStyle",
        fontName=PDF_FONT,
        fontSize=9,
        leading=11,
        alignment=1,
        textColor=colors.HexColor("#94A3B8")
    )

    story.append(Paragraph(title.upper(), title_style))

    headers = [Paragraph("Saat", head_style)] + [Paragraph(day, head_style) for day in DAYS]
    table_data = [headers]

    schedule_map = {(row["day"], row["hour"]): row for row in schedule}

    for hour in HOURS:
        row_cells = [Paragraph(f"<b>{hour}. Saat</b>", cell_main_style)]
        for day in DAYS:
            item = schedule_map.get((day, hour))
            if item:
                main_text = item.get(primary_label) or item.get("lesson_name") or ""
                sub_text = item.get("lesson_name") if primary_label == "class_name" else item.get("teacher_name") or ""
                room_text = f" ({item['room_name']})" if item.get("room_name") else ""

                cell_html = f"<b>{main_text}</b><br/><font size=7.5 color='#475569'>{sub_text}{room_text}</font>"
                row_cells.append(Paragraph(cell_html, cell_main_style))
            else:
                row_cells.append(Paragraph("-", cell_empty_style))
        table_data.append(row_cells)

    col_widths = [50, 145, 145, 145, 145, 145]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#334155')),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F1F5F9')),
        ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    story.append(t)
    doc.build(story)
    return buffer.getvalue()


